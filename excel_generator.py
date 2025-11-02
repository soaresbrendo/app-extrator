import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def gerar_excel(dados_por_cnpj_pedido):
    """
    Gera arquivo Excel com abas por CNPJ e Pedido
    Entrada: {cnpj: {pedido: [produtos]}}
    """
    try:
        output = io.BytesIO()
        wb = Workbook()
        
        # Remover a planilha padrão
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        for cnpj, pedidos in dados_por_cnpj_pedido.items():
            for pedido, produtos in pedidos.items():
                if not produtos:
                    continue
                
                # Formato: 31838128000314-7202309
                cnpj_limpo = cnpj.replace('.', '').replace('/', '').replace('-', '')
                aba_nome = f"{cnpj_limpo}-{pedido}"
                
                # Limitar a 31 caracteres (limite do Excel)
                aba_nome = aba_nome[:31]
                
                # Criar nova aba
                ws = wb.create_sheet(title=aba_nome)
                
                # Obter cabeçalhos das colunas
                if produtos:
                    headers = list(produtos[0].keys())
                    
                    # Escrever cabeçalhos
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.value = header
                        # Negrito para cabeçalhos
                        cell.font = cell.font.copy(bold=True)
                    
                    # Escrever dados
                    for row_idx, produto in enumerate(produtos, start=2):
                        for col_idx, header in enumerate(headers, start=1):
                            ws.cell(row=row_idx, column=col_idx).value = produto.get(header, '')
                    
                    # Ajustar largura das colunas
                    for col_idx, header in enumerate(headers, start=1):
                        max_length = len(str(header))
                        
                        # Verificar comprimento dos dados
                        for row_idx in range(2, len(produtos) + 2):
                            cell_value = ws.cell(row=row_idx, column=col_idx).value
                            if cell_value:
                                max_length = max(max_length, len(str(cell_value)))
                        
                        # Definir largura (mínimo 10, máximo 50)
                        adjusted_width = min(max(max_length + 3, 10), 50)
                        column_letter = get_column_letter(col_idx)
                        ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar no BytesIO
        wb.save(output)
        output.seek(0)
        return output
        
    except Exception as e:
        print(f"Erro ao criar Excel: {e}")
        import traceback
        traceback.print_exc()
        return None
